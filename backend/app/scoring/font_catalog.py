"""
scoring/font_catalog.py — Monotype font catalog lookup.

Loads the pre-built font_catalog.json (generated from
Sales_Inventory_Report_4.23.26.xlsx) and provides O(1) source lookups by
normalized PostScript or CSS family name.

Source categories in the catalog:
  • MT Owned                          – Monotype-owned; fully licensed
  • UFDA                              – Under Font Distribution Agreement; licensed
  • 3rd Party Font                    – Licensed third-party; compliant
  • 3rd Party Font - Unclear PostScript Name – Licensed but name ambiguous
  • Free Fonts                        – Open-source / free; generally compliant
  • In The Wild                       – Unclear ownership; legal review required
"""

from __future__ import annotations

import json
import re
from functools import lru_cache
from pathlib import Path

# ── Catalog paths ──────────────────────────────────────────────────────────────
_CATALOG_JSON = Path(__file__).parent.parent / "font_catalog.json"
_CATALOG_XLSX = Path(__file__).parent.parent / "Sales_Inventory_Report_4.23.26.xlsx"

# Source value that requires legal review
IN_THE_WILD = "In The Wild"

# Sources considered fully licensed / compliant
LICENSED_SOURCES: frozenset[str] = frozenset({
    "MT Owned",
    "UFDA",
    "3rd Party Font",
    "3rd Party Font - Unclear PostScript Name",
    "Free Fonts",
})


def _normalize(name: str) -> str:
    """Collapse a font name to lowercase alphanumeric for fuzzy matching."""
    return re.sub(r"[^a-z0-9]", "", name.lower())


@lru_cache(maxsize=1)
def _load_catalog() -> dict[str, str]:
    """
    Load the catalog once and cache it for the process lifetime.

    Tries the pre-built JSON first (fast, ~0.5 s).
    Falls back to parsing the raw Excel if JSON is absent.
    """
    if _CATALOG_JSON.exists():
        with open(_CATALOG_JSON, "r", encoding="utf-8") as fh:
            return json.load(fh)

    # Fallback: parse Excel (slow – only happens when JSON is missing)
    try:
        import openpyxl  # noqa: PLC0415
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl is required to build the font catalog from Excel. "
            "Run: pip install openpyxl"
        ) from exc

    wb = openpyxl.load_workbook(str(_CATALOG_XLSX), read_only=True)
    ws = wb.worksheets[0]
    headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]

    src_idx    = headers.index("Source")
    ps_idx     = headers.index("Postscript Name")
    family_idx = headers.index("Family Name")

    catalog: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        source = row[src_idx]
        if not source:
            continue
        for raw in (row[ps_idx], row[family_idx]):
            if raw:
                key = _normalize(str(raw))
                if key and key not in catalog:
                    catalog[key] = source
    wb.close()
    return catalog


# ── Public API ─────────────────────────────────────────────────────────────────

def lookup_source(font_name: str) -> str | None:
    """
    Return the catalog source for *font_name*, or None if not found.

    Matching is done on the normalized (lowercase, alphanumeric-only) form so
    that CSS family names like "Arial Hebrew Desk Interface" correctly match the
    PostScript name "ArialHebrewDeskInterface".
    """
    return _load_catalog().get(_normalize(font_name))


def classify(font_name: str) -> str:
    """
    Return one of:
      "in_the_wild"  – source is "In The Wild"; customer must check with legal
      "licensed"     – source is a known compliant category
      "unknown"      – not found in the catalog at all
    """
    source = lookup_source(font_name)
    if source is None:
        return "unknown"
    if source == IN_THE_WILD:
        return "in_the_wild"
    return "licensed"


def warm_cache() -> None:
    """Pre-load the catalog into memory (call at app startup)."""
    _load_catalog()
